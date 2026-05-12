from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Iterable

from openpyxl import load_workbook


REQUIRED_COLUMNS = (
    "phase",
    "prompt_id",
    "theme",
    "subpremise_id",
    "premise_text",
    "suggestion_text",
    "suggestion_model",
    "suggestion_generated_at",
    "generation_prompt_version",
)

OPTIONAL_COLUMNS = (
    "difficulty",
    "condition_slot",
    "participant_slot",
    "notes",
)

ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS


@dataclass(frozen=True)
class MaterialValidation:
    ok: bool
    errors: list[str]
    warnings: list[str]
    counts: dict[str, int]
    rows: list[dict[str, str]]


def parse_material_file(filename: str, content: bytes) -> list[dict[str, str]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig")
        return _rows_from_csv(text)
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return _rows_from_xlsx(content)
    raise ValueError("Unsupported material format. Use .csv or .xlsx.")


def validate_material_rows(rows: Iterable[dict[str, object]]) -> MaterialValidation:
    normalized = [_normalize_row(row) for row in rows]
    errors: list[str] = []
    warnings: list[str] = []
    counts = {"practice": 0, "formal": 0}

    if not normalized:
        errors.append("Material table is empty.")
        return MaterialValidation(False, errors, warnings, counts, [])

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in normalized[0]]
    if missing_columns:
        errors.append("Missing required columns: " + ", ".join(missing_columns))

    seen: set[str] = set()
    for index, row in enumerate(normalized, start=2):
        prompt_id = row.get("prompt_id", "")
        if not prompt_id:
            errors.append(f"Row {index}: prompt_id is required.")
        elif prompt_id in seen:
            errors.append(f"Duplicate prompt_id: {prompt_id}.")
        seen.add(prompt_id)

        phase = row.get("phase", "").lower()
        if phase not in counts:
            errors.append(f"Row {index}: phase must be practice or formal.")
        else:
            counts[phase] += 1

        for column in REQUIRED_COLUMNS:
            if not row.get(column, ""):
                errors.append(f"Row {index}: {column} is required.")

    if counts["practice"] < 5:
        errors.append("At least 5 practice materials are required.")
    if counts["formal"] < 20:
        errors.append("At least 20 formal materials are required.")

    extra_columns = sorted(set(normalized[0]) - set(ALL_COLUMNS))
    if extra_columns:
        warnings.append("Ignoring extra columns: " + ", ".join(extra_columns))

    return MaterialValidation(not errors, errors, warnings, counts, normalized)


def _rows_from_csv(text: str) -> list[dict[str, str]]:
    reader = csv.DictReader(StringIO(text))
    return [dict(row) for row in reader]


def _rows_from_xlsx(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    output: list[dict[str, str]] = []
    for row in rows[1:]:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        output.append({headers[index]: _cell_to_text(value) for index, value in enumerate(row) if index < len(headers)})
    return output


def _normalize_row(row: dict[str, object]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in row.items():
        if key is None:
            continue
        normalized[str(key).strip()] = _cell_to_text(value)
    if "phase" in normalized:
        normalized["phase"] = normalized["phase"].strip().lower()
    return normalized


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
